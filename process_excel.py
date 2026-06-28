#!/usr/bin/env python3
"""
公众号文章列表.xlsx → articles.json
=======================================
运行：python3 process_excel.py

处理逻辑：
- 新加坡生活 → 独立分类，保留子分类
- 东南亚游记 → 游记 + 根据子分类推断文化/美食/生活/商业/交通
- 国内旅行   → 游记 + 根据子分类推断
- 美食/文化/游记/生活/商业/交通 → 保留，追加推断
- 子分类：合并重复（佛教→南传佛教，印度教→印度教文化），删除主分类词
"""

import json
from pathlib import Path
from collections import Counter
import openpyxl

EXCEL_PATH = Path(__file__).parent / "公众号文章列表.xlsx"
JSON_PATH  = Path(__file__).parent / "articles.json"

# ── 子分类标准化 ─────────────────────────────────────────
SUBTAG_REMAP = {
    "佛教":    "南传佛教",
    "佛教文化": "南传佛教",
    "印度教":  "印度教文化",
}
# 这些词是主分类，不应出现在子分类
MAIN_THEME_WORDS = {"文化", "美食", "生活", "商业", "游记", "交通"}

# ── 主题推断关键词 ────────────────────────────────────────
CULTURE_KEYS = {
    "历史", "博物馆", "峇峇娘惹", "华人文化", "建筑", "戏剧", "电影",
    "文学", "讲座", "庆典", "节日", "世界文化遗产", "南传佛教", "佛教",
    "佛教文化", "印度教", "印度教文化", "印度文化", "伊斯兰文化",
    "马来文化", "土著文化", "基督教", "天主教", "欧亚裔文化", "汉丽宝",
    "电影节", "文学节", "展览", "演出", "名人", "战争", "书店", "书",
    "印尼传统文化", "传统文化", "南洋文化",
}
FOOD_KEYS      = {"华人美食", "娘惹菜", "咖啡", "零食", "米其林",
                  "新加坡美食", "马来美食", "印尼美食", "越南美食",
                  "老挝美食", "菲律宾美食", "文莱美食", "中国美食",
                  "韩国美食", "印度美食"}
TRANSPORT_KEYS = {"交通"}
COMMERCE_KEYS  = {"商业"}
LIFE_KEYS      = {"生活"}


def parse_csv(raw):
    """支持中英文逗号的逗号分隔解析"""
    if not raw:
        return []
    return [t.strip() for t in str(raw).replace("，", ",").split(",") if t.strip()]


def parse_country_city(raw):
    """'马来西亚，槟城' → ('马来西亚', '槟城')"""
    parts = parse_csv(raw)
    if not parts:
        return "", ""
    return parts[0], "，".join(parts[1:]) if len(parts) > 1 else ""


def normalize_subtags(raw):
    """标准化子分类：重映射 + 去掉主分类词 + 去重"""
    tags = parse_csv(raw)
    seen = set()
    result = []
    for t in tags:
        t = SUBTAG_REMAP.get(t, t)        # 合并重复
        if t in MAIN_THEME_WORDS:          # 跳过主分类词
            continue
        if t and t not in seen:
            seen.add(t)
            result.append(t)
    return result


def raw_subtag_set(raw):
    """原始子分类集合（含主分类词，用于主题推断）"""
    tags = parse_csv(raw)
    return {SUBTAG_REMAP.get(t, t) for t in tags if t.strip()}


def infer_extra_themes(subtags, current_themes):
    """根据子分类推断额外主题（不覆盖已有）"""
    s = set(subtags)
    themes = list(current_themes)

    def add(theme):
        if theme not in themes:
            themes.append(theme)

    if s & CULTURE_KEYS:   add("文化")
    if s & FOOD_KEYS:      add("美食")
    if s & TRANSPORT_KEYS: add("交通")
    if s & COMMERCE_KEYS:  add("商业")
    if s & LIFE_KEYS:      add("生活")
    return themes


def resolve_themes(tag_raw, subtags):
    """
    分类标签 → 标准化主题列表（可多个）

    新加坡生活            → ["新加坡生活"]（独立分类，不推断）
    东南亚游记/国内旅行    → ["游记"] + 推断
    美食/文化/游记/生活等  → 保留 + 推断
    """
    raw = parse_csv(tag_raw)

    # 含新加坡生活 → 独立分类
    if "新加坡生活" in raw:
        return ["新加坡生活"]

    base = []
    for t in raw:
        if t in ("东南亚游记", "国内旅行", "游记"):
            if "游记" not in base:
                base.append("游记")
        elif t in ("美食", "文化", "生活", "商业", "交通"):
            if t not in base:
                base.append(t)
        else:
            # 未知标签：默认游记
            if "游记" not in base:
                base.append("游记")

    if not base:
        base = ["游记"]

    return infer_extra_themes(subtags, base)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    articles = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 2).value
        if not title:
            continue

        pub_date    = ws.cell(r, 3).value or ""
        collection  = ws.cell(r, 4).value or ""
        country_raw = ws.cell(r, 5).value
        tag_raw     = ws.cell(r, 6).value
        subtag_raw  = ws.cell(r, 7).value
        featured    = bool(ws.cell(r, 8).value)
        rep         = bool(ws.cell(r, 9).value)
        cover_img   = ws.cell(r, 10).value or ""
        url         = ws.cell(r, 11).value or ""

        country, city = parse_country_city(country_raw)
        subtags = normalize_subtags(subtag_raw)
        # 推断主题用原始集合（包含"交通"等主分类词）
        themes  = resolve_themes(tag_raw, raw_subtag_set(subtag_raw))

        articles.append({
            "url":            url,
            "title":          str(title),
            "pub_date":       str(pub_date),
            "cover_img":      str(cover_img),
            "collection":     str(collection),
            "themes":         themes,
            "subtags":        subtags,
            "country":        country,
            "city":           city,
            "featured":       featured,
            "representative": rep,
        })

    JSON_PATH.write_text(
        json.dumps(articles, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"✅ articles.json 已生成：{len(articles)} 篇\n")

    # ── 统计 ──────────────────────────────────────────────
    theme_cnt = Counter()
    country_cnt = Counter()
    subtag_cnt = Counter()
    for a in articles:
        for t in a["themes"]:
            theme_cnt[t] += 1
        if a["country"]:
            country_cnt[a["country"]] += 1
        for s in a["subtags"]:
            subtag_cnt[s] += 1

    print("主题分布：")
    for k, v in theme_cnt.most_common():
        print(f"  {k}: {v}篇")

    print("\n国家 Top10：")
    for k, v in country_cnt.most_common(10):
        print(f"  {k}: {v}篇")

    print("\n子分类 Top15：")
    for k, v in subtag_cnt.most_common(15):
        print(f"  {k}: {v}篇")

    featured_cnt = sum(1 for a in articles if a["featured"])
    rep_cnt      = sum(1 for a in articles if a["representative"])
    print(f"\n精选(10万+): {featured_cnt}篇  |  国家代表作: {rep_cnt}篇")


if __name__ == "__main__":
    main()
